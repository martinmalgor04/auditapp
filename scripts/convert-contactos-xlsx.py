#!/usr/bin/env python3
"""Convierte los xlsx de contactos comerciales a CSVs de seed.

Uso:
    python3 scripts/convert-contactos-xlsx.py "<Lista de Contactos.xlsx>" "<CLIENTES POTENCIALES (Respuestas).xlsx>"

Genera:
    seed/clientes-tango.csv   — clientes fijos Tango con datos de licencia
    seed/prospectos.csv       — prospectos relevados (formulario + base de contactos)

Los ids son UUIDv5 determinísticos derivados de la razón social, así el seed
es idempotente y re-correr el conversor con data actualizada conserva los ids.
Los clientes Tango que ya existen en clientes-presupuestossys.csv (match por
razón social normalizada) reusan el id existente para no duplicar filas.

Requiere: openpyxl (pip install openpyxl)
"""

import csv
import re
import sys
import unicodedata
import uuid
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SEED_DIR = ROOT / 'seed'
PRESUPUESTOS_CSV = SEED_DIR / 'clientes-presupuestossys.csv'

NAMESPACE = uuid.uuid5(uuid.NAMESPACE_DNS, 'auditapp.serviciosysistemas.com.ar')

PLACEHOLDERS = {'BUSCAR', 'CONSULTAR', 'CONSULTR', 'VER', 'N/A', '-'}


def norm_key(name: str) -> str:
    """Normaliza razón social para matching: sin acentos ni puntuación, mayúsculas."""
    s = unicodedata.normalize('NFKD', name)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'[^A-Z0-9]', '', s.upper())


def clean(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return '' if text.upper() in PLACEHOLDERS else text


def clean_phone(value) -> str:
    return clean(value)


def clean_mail(value) -> str:
    text = clean(value)
    return text if '@' in text else ''


def clean_date(value) -> str:
    """'2025-01-23 12:00:00.000' o datetime → '2025-01-23'."""
    text = clean(value)
    match = re.match(r'(\d{4}-\d{2}-\d{2})', text)
    return match.group(1) if match else ''


def load_existing_ids() -> dict[str, str]:
    ids: dict[str, str] = {}
    with open(PRESUPUESTOS_CSV, newline='', encoding='utf-8') as fh:
        for row in csv.DictReader(fh):
            ids.setdefault(norm_key(row['razon_social']), row['id'])
    return ids


def stable_id(prefix: str, razon_social: str, existing: dict[str, str]) -> tuple[str, bool]:
    key = norm_key(razon_social)
    if key in existing:
        return existing[key], True
    return str(uuid.uuid5(NAMESPACE, f'{prefix}:{key}')), False


def convert_tango(xlsx_path: str, existing: dict[str, str]) -> int:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Pag. 0']
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]

    out_path = SEED_DIR / 'clientes-tango.csv'
    matched = 0
    with open(out_path, 'w', newline='', encoding='utf-8') as fh:
        writer = csv.writer(fh)
        writer.writerow([
            'id', 'razon_social', 'contacto', 'telefono', 'email', 'tipo',
            'terminales', 'venc_escala', 'version', 'version_detectada',
            'lic_categoria', 'sueldos', 'motivo'
        ])
        for r in rows:
            razon = clean(r[0])
            if not razon:
                continue
            cid, was_match = stable_id('tango', razon, existing)
            matched += was_match
            telefono = clean_phone(r[2])
            email = clean_mail(r[3])
            # a veces cargaron un teléfono en la columna de mail
            if not email and not telefono:
                telefono = clean_phone(r[3]) if re.fullmatch(r'\d{7,}', clean(r[3])) else ''
            writer.writerow([
                cid, razon, clean(r[1]), telefono, email, clean(r[5]),
                clean(r[6]), clean_date(r[7]), clean(r[9]), clean(r[13]),
                clean(r[10]), clean(r[12]), clean(r[11])
            ])
    print(f'clientes-tango.csv: {len(rows)} filas ({matched} matcheadas con presupuestos)')
    return len(rows)


def convert_prospectos(xlsx_path: str, existing: dict[str, str]) -> int:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    merged: dict[str, list] = {}

    def add(row: list) -> None:
        """Mergea por id: primer valor no vacío gana, observaciones se concatenan."""
        prev = merged.get(row[0])
        if prev is None:
            merged[row[0]] = row
            return
        for i, value in enumerate(row):
            if i == 10 and value and prev[10] and value not in prev[10]:
                prev[10] = f'{prev[10]} | {value}'
            elif not prev[i]:
                prev[i] = value

    # Hoja 1: respuestas del formulario de relevamiento en calle
    ws = wb['Respuestas de formulario 1']
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]
    last: list | None = None
    for r in rows:
        razon = clean(r[1])
        obs = clean(r[9])
        if not razon:
            # fila sin razón social = continuación de la observación anterior
            if last is not None and obs:
                last[10] = (last[10] + ' ' + obs).strip()
            continue
        cid, _ = stable_id('prospecto', razon, existing)
        relevado = r[0].strftime('%Y-%m-%d %H:%M:%S-03') if r[0] else ''
        dolor = clean(r[6])
        if dolor:
            obs = f'Dolor: {dolor}. {obs}'.strip()
        last = [
            cid, razon, clean(r[2]), clean_phone(r[3]), clean_mail(r[4]),
            clean(r[5]), '', '', clean(r[7]), clean(r[8]), obs,
            relevado, 'formulario'
        ]
        add(last)

    # Hoja 2: base de contactos sueltos
    ws2 = wb['BASE DE CONTACTOS']
    for r in [r for r in ws2.iter_rows(min_row=2, values_only=True) if any(r)]:
        razon = clean(r[2])
        if not razon:
            continue
        cid, _ = stable_id('prospecto', razon, existing)
        referente = ' '.join(p for p in (clean(r[0]), clean(r[1])) if p)
        add([
            cid, razon, referente, clean_phone(r[4]), clean_mail(r[3]),
            '', clean(r[6]).rstrip('\\'), clean(r[5]), '', '', '',
            '', 'base_contactos'
        ])

    out_path = SEED_DIR / 'prospectos.csv'
    with open(out_path, 'w', newline='', encoding='utf-8') as fh:
        writer = csv.writer(fh)
        writer.writerow([
            'id', 'razon_social', 'referente', 'telefono', 'email', 'direccion',
            'rubro', 'pagina', 'tiene_software', 'nivel_interes', 'observaciones',
            'relevado_at', 'fuente'
        ])
        writer.writerows(merged.values())
    print(f'prospectos.csv: {len(merged)} filas')
    return len(merged)


def main() -> None:
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    existing = load_existing_ids()
    convert_tango(sys.argv[1], existing)
    convert_prospectos(sys.argv[2], existing)


if __name__ == '__main__':
    main()
